## PAQUETERIAS ##
import os, re, warnings, pickle
import numpy as np
import pandas as pd
from scipy.optimize   import minimize_scalar, differential_evolution
from scipy.stats      import weibull_min, gaussian_kde, kstest
from sklearn.metrics  import r2_score
import openpyxl

warnings.filterwarnings("ignore")

## CONFIGURACIÓN ##
K_CEROS  = 10
EPSILON  = 0.01
THETAS   = [30.0, 35.0, 40.0]   ## Umbrales para P(X > θ)

## Lista de datasets: (etiqueta_ciudad, etiqueta_periodo, ruta_archivo)
DATASETS = [
    ("CDMX",        "1940-1968", "data/data_cdmx_1.xlsx"),
    ("CDMX",        "1969-1996", "data/data_cdmx_2.xlsx"),
    ("CDMX",        "1997-2026", "data/data_cdmx_3.xlsx"),
    ("Chihuahua",   "1940-1968", "data/data_chh_1.xlsx"),
    ("Chihuahua",   "1969-1996", "data/data_chh_2.xlsx"),
    ("Chihuahua",   "1997-2026", "data/data_chh_3.xlsx"),
    ("Guadalajara", "1940-1968", "data/data_guad_1.xlsx"),
    ("Guadalajara", "1969-1996", "data/data_guad_2.xlsx"),
    ("Guadalajara", "1997-2026", "data/data_guad_3.xlsx"),
    ("Monterrey",   "1940-1968", "data/data_mty_1.xlsx"),
    ("Monterrey",   "1969-1996", "data/data_mty_2.xlsx"),
    ("Monterrey",   "1997-2026", "data/data_mty_3.xlsx"),
    ("Phoenix",     "1940-1968", "data/data_phnx_1.xlsx"),
    ("Phoenix",     "1969-1996", "data/data_phnx_2.xlsx"),
    ("Phoenix",     "1997-2026", "data/data_phnx_3.xlsx"),
]

## UTILIDADES ##
def sig_i(carpeta="output", prefijo="base_"):
    """Devuelve la ruta de archivo Excel de salida con el siguiente índice."""
    os.makedirs(carpeta, exist_ok=True)
    nums = []
    patron = rf"{prefijo}(\d+)\.xlsx$"
    for a in os.listdir(carpeta):
        match = re.match(patron, a)
        if match:
            nums.append(int(match.group(1)))
    return os.path.join(carpeta, f"{prefijo}{(max(nums) + 1 if nums else 1)}.xlsx")


def freedman_diaconis_bin(data):
    """Ancho de bin según regla de Freedman-Diaconis."""
    q75, q25 = np.percentile(data, [75, 25])
    iqr = q75 - q25
    if iqr == 0:
        raise ValueError("IQR = 0.")
    return 2.0 * iqr / (len(data) ** (1.0 / 3.0))


def calcular_mu_star(tmax, n_grid=2000):
    """μ* = pico de la densidad KDE gaussiana con regla de Scott."""
    sigma_hat = np.std(tmax, ddof=1)
    h_scott   = 1.059 * sigma_hat * len(tmax) ** (-1.0 / 5.0)
    kde       = gaussian_kde(tmax, bw_method=h_scott / sigma_hat)
    x_grid    = np.linspace(tmax.min(), tmax.max(), n_grid)
    densidad  = kde(x_grid)
    mu_star   = float(x_grid[np.argmax(densidad)])
    return mu_star, h_scott, kde, x_grid, densidad


## WLR
def compute_wlr(t_pos, A, n_pos, weight):
    """
    Regresión lineal ponderada Weibull sobre clases positivas.
    Posiciones: F_i = (i-A)/(n_pos+1-2A),  A ∈ [0, 0.5)
    """
    i_ranks = np.arange(1, n_pos + 1, dtype=float)
    denom   = n_pos + 1.0 - 2.0 * A
    if denom <= 0:
        return None
    F = (i_ranks - A) / denom
    if np.any(F <= 0) or np.any(F >= 1):
        return None

    X     = np.log(t_pos)
    omF   = 1.0 - F
    lnomF = np.log(omF)
    Y     = np.log(-lnomF)

    W = (omF * lnomF)**2 if weight == 'A' \
        else n_pos * omF * lnomF**2 / F

    Sw = W.sum()
    if Sw == 0:
        return None
    Xm = (X * W).sum() / Sw
    Ym = (Y * W).sum() / Sw
    dx = X - Xm;  dy = Y - Ym
    Sxy = (W * dx * dy).sum()
    Sxx = (W * dx * dx).sum()
    Syy = (W * dy * dy).sum()
    if Sxx == 0 or Sxy == 0 or Syy == 0:
        return None

    k         = Sxy / Sxx               ## shape = pendiente WLR
    intercept = Ym - k * Xm
    if k <= 0:
        return None
    lam = np.exp(-intercept / k)     
    R2  = (Sxy**2) / (Sxx * Syy)

    return dict(k=k, lam=lam, intercept=intercept, R2_wlr=R2,
                F=F, X=X, Y=Y, W=W, Xm=Xm, Ym=Ym,
                Sxy=Sxy, Sxx=Sxx, Syy=Syy)


## WEIBULL Y MODELO DE MEZCLA ##
def wdist(t, k, lam):
    """F_W(t; k, λ) = 1 − exp(−(t/λ)^k)."""
    return weibull_min.cdf(t, c=k, scale=lam)


def weibull_mix(t_vals, C_vals, k, lam):
    """
    FDA del modelo de mezcla:
      U[0]   = p0 = C_vals[0]                    (clase cero: exacto)
      U[i>0] = p0 + (1−p0)·F_W(t_i; k, λ)       (clases positivas)
    """
    p0   = float(C_vals[0])
    U    = np.empty_like(C_vals)
    U[0] = p0
    U[1:] = p0 + (1.0 - p0) * wdist(t_vals[1:], k, lam)
    diff  = C_vals - U
    return U, diff, diff**2


def r2_mix(C_vals, U_vals):
    """R² = 1 − SS_res/SS_tot  (sklearn r2_score, puede ser negativo)."""
    if len(C_vals) < 2:
        return 0.0
    return float(r2_score(C_vals, U_vals))


## DIAGNÓSTICOS Y EXCEDENCIA
def diagnosticos_ajuste(T_pos, k, lam):

    if len(T_pos) < 5 or k <= 0 or lam <= 0:
        return dict(KS_stat=np.nan, KS_pval=np.nan, AD_stat=np.nan)
    ks_stat, ks_pval = kstest(T_pos, 'weibull_min', args=(k, 0, lam))
    try:
        n  = len(T_pos)
        u  = weibull_min.cdf(np.sort(T_pos), k, 0, lam)
        u  = np.clip(u, 1e-12, 1 - 1e-12)
        ad = -n - np.sum((2*np.arange(1, n+1) - 1) / n
                         * (np.log(u) + np.log(1 - u[::-1])))
    except Exception:
        ad = np.nan
    return dict(KS_stat=ks_stat, KS_pval=ks_pval, AD_stat=ad)


def prob_excede(theta, mu_star, p0, k, lam, tmax_raw):
    """
    P(X > θ):
      θ > μ*  →  (1−p0)·exp(−((θ−μ*)/λ)^k)   [Weibull]
      θ = μ*  →  1 − p0
      θ < μ*  →  CDF empírica
    """
    if np.isclose(theta, mu_star, atol=1e-6):
        return float(1.0 - p0), "exacto: 1−p0"
    elif theta > mu_star:
        val = float((1.0 - p0) * np.exp(-((theta - mu_star) / lam) ** k))
        return val, "Weibull"
    else:
        return float(np.mean(tmax_raw > theta)), "empírica"


## PREPARAR CLASES ##
def preparar_clases(t_raw, f_raw, total):

    idx = np.argsort(t_raw)
    t_v = t_raw[idx].copy();  f_v = f_raw[idx].copy()
    C_v = np.cumsum(f_v) / total
    lim = next((i for i, c in enumerate(C_v) if c >= 1.0), len(C_v))
    t_v, f_v, C_v = t_v[:lim], f_v[:lim], C_v[:lim]
    return t_v, f_v, C_v, np.arange(1, lim + 1, dtype=float)


## BUCLE PRINCIPAL ##
os.makedirs("output", exist_ok=True)
ALL_RESULTS = {}   ## {ciudad: {periodo: {peso: result_dict}}}

for ciudad, periodo, filepath in DATASETS:

    if not os.path.exists(filepath):
        print(f"\n No encontrado: {filepath} , saltando")
        continue

    print(f"\n{'-'*60}")
    print(f"  CIUDAD: {ciudad}   PERIODO: {periodo}")
    print(f"{'-'*60}")

    df   = pd.read_excel(filepath)
    tmax = pd.to_numeric(df["i_tmax"], errors="coerce").dropna().to_numpy(dtype=float)
    if len(tmax) < 30:
        print(f"Pocos datos ({len(tmax)}) , saltando.")
        continue

    ## μ* via KDE ##
    mu_star, h_scott, kde_obj, x_grid, densidad = calcular_mu_star(tmax)
    mediana = float(np.median(tmax))
    print(f"\n   μ* (umbral modal KDE)")
    print(f"     Mediana           = {mediana:.4f} °C")
    print(f"     μ*                = {mu_star:.4f} °C")
    print(f"     Ancho banda Scott = {h_scott:.4f} °C")
    print(f"     Diferencia        = {mu_star - mediana:+.4f} °C")

    ## T y clases ##
    T = np.clip(tmax - mu_star, 0, None)
    try:
        ancho_bin = freedman_diaconis_bin(T)
    except ValueError:
        ancho_bin = (T.max() - T.min()) / max(1, int(np.sqrt(len(T))))

    t_raw = [EPSILON];  f_raw = [int(np.sum(T == 0))]
    j, zc = 1, 0
    while True:
        li, ls = (j-1)*ancho_bin, j*ancho_bin
        fq = int(np.sum((T > li) & (T <= ls)))
        t_raw.append(j*ancho_bin);  f_raw.append(fq)
        zc = zc + 1 if fq == 0 else 0
        if zc >= K_CEROS:
            break
        j += 1
    while len(f_raw) > 1 and f_raw[-1] == 0:
        t_raw.pop();  f_raw.pop()

    t_raw = np.array(t_raw, dtype=float)
    f_raw = np.array(f_raw, dtype=float)
    total = f_raw.sum()
    t_v, f_v, C_v, i_v = preparar_clases(t_raw, f_raw, total)
    n     = len(t_v)
    t_pos = t_v[1:]         ## clases POSITIVAS para el WLR
    n_pos = len(t_pos)
    p0    = float(C_v[0])   ## masa en cero

    print(f"\n  n obs={len(tmax)}  T=0={int(np.sum(T==0))} ({p0*100:.1f}%)"
          f"  T>0={int(np.sum(T>0))}  bin={ancho_bin:.4f}  clases WLR={n_pos}")

    ## MLE como punto de partida para T y clases ##
    T_pos   = T[T > 0]
    k_mle, _, lam_mle = weibull_min.fit(T_pos, floc=0)
    U_mle, _, _ = weibull_mix(t_v, C_v, k_mle, lam_mle)
    R2_mle = r2_mix(C_v, U_mle)
    print(f"  MLE: k={k_mle:.4f}  λ={lam_mle:.4f}  R²={R2_mle:.4f}")

    if ciudad not in ALL_RESULTS:
        ALL_RESULTS[ciudad] = {}

    periodo_pesos = {}

    for peso, label in [('A', 'Peso A (Bergman)')]:

        print(f"\n  {'─'*55}")
        print(f"  {label}")
        print(f"  {'─'*55}")

        ## WLR inicial A=0
        A_INI   = 0.0
        r0      = compute_wlr(t_pos, A_INI, n_pos, peso)
        if r0 is None:
            r0 = dict(k=k_mle, lam=lam_mle, R2_wlr=0.0)
        print(f"  [WLR A=0]  k={r0['k']:.4f}  λ={r0['lam']:.4f}"
              f"  R²_WLR={r0['R2_wlr']:.4f}")
        U0, _, _ = weibull_mix(t_v, C_v, r0['k'], r0['lam'])
        print(f"             R²_mezcla={r2_mix(C_v, U0):.4f}")

        ## Optimizar A
        def obj_A(A_val, tp=t_pos, np_=n_pos, wt=peso, tv=t_v, Cv=C_v):

            r = compute_wlr(tp, A_val, np_, wt)
            if r is None:
                return 1.0
            U, _, _ = weibull_mix(tv, Cv, r['k'], r['lam'])
            return 1.0 - r2_mix(Cv, U)

        res_A = minimize_scalar(obj_A, bounds=(0.0, 0.5 - 1e-9),
                                method='bounded', options={'xatol': 1e-10})
        A_opt = float(res_A.x)

        ## WLR con A óptimo
        r_wlr = compute_wlr(t_pos, A_opt, n_pos, peso)
        if r_wlr is None:
            r_wlr = r0
        k_wlr, lam_wlr = r_wlr['k'], r_wlr['lam']
        U_wlr, d_wlr, d2_wlr = weibull_mix(t_v, C_v, k_wlr, lam_wlr)
        R2_wlr = r2_mix(C_v, U_wlr)

        ref = ("Weibull" if A_opt < 0.15 else
               "Benard"  if A_opt < 0.34 else
               "Blom"    if A_opt < 0.44 else "Hazen")
        print(f"  [A_opt={A_opt:.5f} ≈ {ref}]  "
              f"k={k_wlr:.4f}  λ={lam_wlr:.4f}  R²_mezcla={R2_wlr:.4f}")

        ## Optimización conjunta (k, λ) — differential_evolution
        print(f"  [Opt. conjunta (k,λ)] differential_evolution...", flush=True)

        def obj_kl(params, tv=t_v, Cv=C_v):
            k_, l_ = params
            if k_ <= 0 or l_ <= 0:
                return 2.0
            U, _, _ = weibull_mix(tv, Cv, k_, l_)
            return 1.0 - r2_mix(Cv, U)

        k_lo  = max(0.05, min(k_mle, k_wlr) * 0.1)
        k_hi  = max(k_mle, k_wlr) * 10.0
        l_lo  = max(0.001, min(lam_mle, lam_wlr) * 0.01)
        l_hi  = max(lam_mle, lam_wlr) * 100.0

        res_jt = differential_evolution(
            obj_kl, bounds=[(k_lo, k_hi), (l_lo, l_hi)],
            seed=42, maxiter=2000, tol=1e-12,
            popsize=25, mutation=(0.5, 1.8), recombination=0.9,
            workers=1
        )
        k_opt, lam_opt = res_jt.x
        U_opt, d_opt, d2_opt = weibull_mix(t_v, C_v, k_opt, lam_opt)
        R2_opt = r2_mix(C_v, U_opt)
        print(f"  k_opt={k_opt:.6f}  λ_opt={lam_opt:.6f}  R²={R2_opt:.6f}")

        ## Alpha y Beta finales paper
        k_paper  = 1.0 / k_opt
        lam_paper = lam_opt ** (-k_opt)


        ## Diagnósticos
        diag = diagnosticos_ajuste(T_pos, k_opt, lam_opt)
        print(f"  KS={diag['KS_stat']:.4f}  p={diag['KS_pval']:.4f}"
              f"  AD={diag['AD_stat']:.4f}")

        ## Excedencias
        excedencias = {}
        for theta in THETAS:
            val, met = prob_excede(theta, mu_star, p0, k_opt, lam_opt, tmax)
            excedencias[theta] = (val, met)
            print(f"  P(X>{theta}°C) = {val*100:.3f}%  [{met}]")

        periodo_pesos[peso] = dict(
            ciudad=ciudad, periodo=periodo, label=label, peso=peso,
            ## datos brutos (para graficas.py)
            tmax_raw=tmax, T=T, T_pos=T_pos,
            mu_star=mu_star, mediana=mediana,
            h_scott=h_scott,
            x_grid=x_grid, densidad=densidad,
            ## clases
            t_v=t_v, f_v=f_v, C_v=C_v, i_v=i_v,
            n=n, n_pos=n_pos, ancho_bin=ancho_bin, p0=p0,
            ## MLE
            k_mle=k_mle, lam_mle=lam_mle, R2_mle=R2_mle,
            ## WLR
            A_ini=A_INI, A_opt=A_opt,
            r_wlr=r_wlr,
            k_wlr=k_wlr, lam_wlr=lam_wlr,
            R2_wlr_reg=r_wlr['R2_wlr'], R2_wlr_mix=R2_wlr,
            U_wlr=U_wlr, diff_wlr=d_wlr, diff2_wlr=d2_wlr,
            ## óptimo conjunto
            k_opt=k_opt, lam_opt=lam_opt, k_paper=k_paper, lam_paper=lam_paper, R2_opt=R2_opt,
            U_opt=U_opt, diff_opt=d_opt, diff2_opt=d2_opt,
            ## diagnósticos y excedencias
            KS_stat=diag['KS_stat'], KS_pval=diag['KS_pval'],
            AD_stat=diag['AD_stat'],
            excedencias=excedencias,
        )

    ALL_RESULTS[ciudad][periodo] = periodo_pesos


## EXCEL ##
file_output = sig_i()
wb = openpyxl.Workbook()
if wb.active:
    wb.remove(wb.active)

COL = {'A':1,'B':2,'C':3,'D':4,'F':6,'G':7,
       'I':9,'J':10,'K':11,'L':12,'M':13,'N':14,
       'O':15,'P':16,'Q':17,'R':18,'S':19,
       'U':21,'V':22,'W':23,'Y':25,'Z':26,'AA':27}

HEADS = {
    'A':'t_i','B':'Frecuencia','C':'%acumulado','D':'i',
    'F':'Parámetro','G':'Valor',
    'I':'F_i=(i-A)/(n_pos+1-2A)','J':'X=ln(t)','K':'Y=ln(-ln(1-F))',
    'L':'W_i','M':'X·W','N':'Y·W',
    'O':'(Y-Ym)(X-Xm)','P':'(Y-Ym)(X-Xm)·W',
    'Q':'(X-Xm)²','R':'(X-Xm)²·W','S':'(Y-Ym)²·W',
    'U':'U_wlr (mezcla)','V':'diff_wlr','W':'diff²_wlr',
    'Y':'U_opt (mezcla)','Z':'diff_opt','AA':'diff²_opt',
}


def write_sheet(wb, nombre, r):
    'Escribe una hoja de Excel con los resultados de un dataset.'
    ws = wb.create_sheet(nombre[:31])
    for col, lbl in HEADS.items():
        ws.cell(1, COL[col]).value = lbl

    rw = r['r_wlr'] or {}
    W  = rw.get('W', np.zeros(r['n_pos']))
    X  = rw.get('X', np.zeros(r['n_pos']))
    Y  = rw.get('Y', np.zeros(r['n_pos']))
    ws.cell(1, COL['L']).value  = W.sum()
    ws.cell(1, COL['P']).value  = rw.get('Sxy', 0)
    ws.cell(1, COL['R']).value  = rw.get('Sxx', 0)
    ws.cell(1, COL['S']).value  = rw.get('Syy', 0)
    ws.cell(1, COL['W']).value  = r['diff2_wlr'].sum()
    ws.cell(1, COL['AA']).value = r['diff2_opt'].sum()

    params = [
        ('Ciudad',               r['ciudad']),
        ('Periodo',              r['periodo']),
        ('μ* (moda KDE, °C)',   r['mu_star']),
        ('Mediana (ref, °C)',    r['mediana']),
        ('h Scott',              r['h_scott']),
        ('p0 = P(T=0)',          r['p0']),
        ('n clases WLR',         r['n_pos']),
        ('Bin F-D (°C)',         r['ancho_bin']),
        ('',                     ''),
        ('─ MLE (punto inicial) ─', ''),
        ('k_mle',                r['k_mle']),
        ('λ_mle',                r['lam_mle']),
        ('R² mezcla (MLE)',      r['R2_mle']),
        ('',                     ''),
        ('─ WLR + opt A ─',      ''),
        ('A_opt',                r['A_opt']),
        ('k_wlr',                r['k_wlr']),
        ('λ_wlr',                r['lam_wlr']),
        ('R² WLR (regresión)',   r['R2_wlr_reg']),
        ('R² mezcla (WLR)',      r['R2_wlr_mix']),
        ('',                     ''),
        ('─ Óptimo conjunto ─',  ''),
        ('k_opt',                r['k_paper']),
        ('λ_opt',                r['lam_paper']),
        ('R² mezcla FINAL',      r['R2_opt']),
        ('',                     ''),
        ('─ Diagnósticos ─',     ''),
        ('KS stat',              r['KS_stat']),
        ('KS p-valor',           r['KS_pval']),
        ('AD stat',              r['AD_stat']),
    ]
    for theta, (val, met) in r['excedencias'].items():
        params.append((f'P(X>{theta}°C) [{met}]', round(val, 6)))

    for ri, (lbl, val) in enumerate(params, start=2):
        ws.cell(ri, COL['F']).value = lbl
        ws.cell(ri, COL['G']).value = val

    ## Datos por clase
    Xm  = rw.get('Xm', 0);  Ym = rw.get('Ym', 0)
    F_i = np.full(r['n'], np.nan)
    if 'F' in rw:
        F_i[1:1+len(rw['F'])] = rw['F']

    W_ext = np.zeros(r['n']); X_ext = np.zeros(r['n']); Y_ext = np.zeros(r['n'])
    nf = min(r['n']-1, len(W))
    W_ext[1:1+nf] = W[:nf];  X_ext[1:1+nf] = X[:nf];  Y_ext[1:1+nf] = Y[:nf]

    for ri, idx in enumerate(range(r['n']), start=2):
        ws.cell(ri, COL['A']).value  = r['t_v'][idx]
        ws.cell(ri, COL['B']).value  = int(r['f_v'][idx])
        ws.cell(ri, COL['C']).value  = r['C_v'][idx]
        ws.cell(ri, COL['D']).value  = idx + 1
        ws.cell(ri, COL['I']).value  = None if np.isnan(F_i[idx]) else F_i[idx]
        ws.cell(ri, COL['J']).value  = X_ext[idx]
        ws.cell(ri, COL['K']).value  = Y_ext[idx]
        ws.cell(ri, COL['L']).value  = W_ext[idx]
        ws.cell(ri, COL['M']).value  = X_ext[idx]*W_ext[idx]
        ws.cell(ri, COL['N']).value  = Y_ext[idx]*W_ext[idx]
        Xk = X_ext[idx]; Yk = Y_ext[idx]; Wk = W_ext[idx]
        ws.cell(ri, COL['O']).value  = (Yk-Ym)*(Xk-Xm)
        ws.cell(ri, COL['P']).value  = (Yk-Ym)*(Xk-Xm)*Wk
        ws.cell(ri, COL['Q']).value  = (Xk-Xm)**2
        ws.cell(ri, COL['R']).value  = (Xk-Xm)**2*Wk
        ws.cell(ri, COL['S']).value  = (Yk-Ym)**2*Wk
        ws.cell(ri, COL['U']).value  = r['U_wlr'][idx]
        ws.cell(ri, COL['V']).value  = r['diff_wlr'][idx]
        ws.cell(ri, COL['W']).value  = r['diff2_wlr'][idx]
        ws.cell(ri, COL['Y']).value  = r['U_opt'][idx]
        ws.cell(ri, COL['Z']).value  = r['diff_opt'][idx]
        ws.cell(ri, COL['AA']).value = r['diff2_opt'][idx]


for ciudad, periodos in ALL_RESULTS.items():
    for periodo, pesos in periodos.items():
        for peso, r in pesos.items():
            nombre = f"{ciudad[:5]}_{periodo[-9:]}_{peso}"
            write_sheet(wb, nombre, r)

## Hoja RESUMEN
ws_r = wb.create_sheet('RESUMEN')
header = ['Ciudad','Periodo','μ* (°C)','Mediana (°C)',
          'A_opt','k_wlr','λ_wlr','R²_WLR_mix',
          'k_opt','λ_opt','R²_FINAL','KS_stat','KS_pval','AD_stat'] + \
         [f'P(X>{θ}°C)' for θ in THETAS]
for c, h in enumerate(header, 1):
    ws_r.cell(1, c).value = h

ri = 2
for ciudad, periodos in ALL_RESULTS.items():
    for periodo, pesos in periodos.items():
        for peso, r in pesos.items():
            fila = [ciudad, periodo,
                    round(r['mu_star'],4), round(r['mediana'],4),
                    round(r['A_opt'],6),
                    round(r['k_wlr'],6), round(r['lam_wlr'],6),
                    round(r['R2_wlr_mix'],6),
                    round(r['k_paper'],6), round(r['lam_paper'],6),
                    round(r['R2_opt'],6),
                    round(r['KS_stat'],4), round(r['KS_pval'],4),
                    round(r['AD_stat'],4)] + \
                   [round(r['excedencias'].get(θ,(np.nan,))[0],6) for θ in THETAS]
            for c, v in enumerate(fila, 1):
                ws_r.cell(ri, c).value = v
            ri += 1

wb.save(file_output)

## Guardar para graficas.py
with open("output/all_results.pkl", "wb") as f:
    pickle.dump(ALL_RESULTS, f)

print(f"\n{'═'*60}")
print(f"Excel  : {file_output}")
print(f"Pickle : output/all_results.pkl")
print(f"{'═'*60}")